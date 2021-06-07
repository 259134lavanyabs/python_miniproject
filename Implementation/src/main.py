# Python program to read an excel file
  
# import openpyxl module
import openpyxl

# Give the location of the file
path = "C:\\Users\\Admin\\Desktop\\pythonexcel.xlsx"
  
# To open the workbook 
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
  
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active


def my_ps_number(ip_ws):
    ps_nums=[]
    for row in ws.iter_rows(min_row=1, max_col=1, max_row=16, values_only=True):
        ps_nums.append(list(row))
    return ps_nums

def show_ps_number(ps):
    for item in ps:
        print(item)

def main():
    print("\nEnter the PS number from the below list:\n")
    ps = my_ps_number(ws)
    show_ps_number(ps)
    user_choice = int(input("\nEnter the Ps number:"))

main()
# Loop will print all values
# of first column 
m_row = sheet_obj.max_row
for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    print(cell_obj.value)
    
# Will print a particular row value    
max_col = sheet_obj.max_column
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = 2, column = i)
    print(cell_obj.value, end = " ")    
    






